import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Template Import Anggota ABB"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Headers
headers = [
    "Nama Anggota",
    "Kontak (Email / Telp)",
    "Alamat",
    "Jabatan & Chapte",
    "Status",
    "Aksi",
    "NIK",
    "Timestamp",
    "Email Address",
    "Foto Profil Bebas"
]

ws.append(headers)

# Sample rows
sample_rows = [
    [
        "Adipta Yanuardie",
        "08123456789",
        "Jl. Ahmad Yani No. 2, Bekasi",
        "Ketua Umum - Bekasi Chapter",
        "active",
        "valid",
        "3275012345670001",
        "2026-08-11 10:00:00",
        "adipta@abbcommunity.id",
        "https://drive.google.com/open?id=1aBSRn5GMsR8YsXJTgCyqLCS5cMO3ZPPI"
    ],
    [
        "Fatwa",
        "08198765432",
        "Jl. Jendral Sudirman No. 45, Jakarta",
        "Wakil Ketua Umum - Jakarta Chapter",
        "active",
        "valid",
        "3171098765430002",
        "2026-08-11 10:05:00",
        "fatwa@abbcommunity.id",
        "https://drive.google.com/open?id=1aBSRn5GMsR8YsXJTgCyqLCS5cMO3ZPPI"
    ],
    [
        "Robby Viory Fansya",
        "08571234567",
        "Bekasi Chapter Basecamp",
        "Lead Digital & IT Architect - Bekasi Chapter",
        "active",
        "valid",
        "3275056789010003",
        "2026-08-11 10:10:00",
        "obbyvior@gmail.com",
        "https://drive.google.com/open?id=1aBSRn5GMsR8YsXJTgCyqLCS5cMO3ZPPI"
    ]
]

for row in sample_rows:
    ws.append(row)

# Styling
header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Calibri", size=10)
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Apply header styles
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Apply data cell styles & auto column widths
ws.row_dimensions[1].height = 28

for row in ws.iter_rows(min_row=2, max_row=len(sample_rows) + 1, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Save files
output_public = r"c:\Users\Mojo\Desktop\abbcommunity.github.io\public\templates\template_import_anggota_abb.xlsx"
output_docs = r"c:\Users\Mojo\Desktop\abbcommunity.github.io\docs\template_import_anggota_abb.xlsx"

wb.save(output_public)
wb.save(output_docs)
print(f"SUCCESS: Excel template saved to {output_public}")
